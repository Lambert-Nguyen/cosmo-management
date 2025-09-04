"""
Excel Import Service for Booking Schedules

This service handles importing booking schedules from Excel files (specifically the 'cleaning schedule' sheet)
and automatically creates/updates bookings and associated tasks.
"""

# import openpyxl  # Moved to function level to avoid circular imports
from datetime import datetime, time, timedelta
from decimal import Decimal
import logging
from typing import Dict, List, Tuple, Optional, TYPE_CHECKING
from django.db import transaction
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.core.exceptions import ValidationError
import zoneinfo

from api.models import (
    Booking, Property, Task, BookingImportLog, BookingImportTemplate
)

if TYPE_CHECKING:
    from django.contrib.auth.models import AbstractBaseUser
    User = AbstractBaseUser
else:
    User = get_user_model()  # Agent's recommendation: Use get_user_model() for future-proofing

logger = logging.getLogger(__name__)

class ExcelImportService:
    """Service for importing booking schedules from Excel files."""
    
    def __init__(self, user: User, template: Optional[BookingImportTemplate] = None):
        self.user = user
        self.template = template
        self.import_log = None
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.total_rows = 0
    
    def _import_tz(self):
        """Get the appropriate timezone for import operations."""
        try:
            # Try to get user's profile timezone, fallback to Tampa/New York
            return zoneinfo.ZoneInfo(getattr(getattr(self.user, "profile", None), "timezone", "America/New_York"))
        except Exception:
            return zoneinfo.ZoneInfo("America/New_York")
        
    def import_excel_file(self, excel_file, sheet_name: str = 'Cleaning schedule') -> Dict:
        """
        Import bookings from Excel file with proper timezone activation.
        
        Args:
            excel_file: Excel file to import
            sheet_name: Name of the sheet to import from
            
        Returns:
            Dict with import results
        """
        # Agent's recommendation: Wrap the import in timezone override
        with timezone.override(self._import_tz()):
            return self._import_excel_file_inner(excel_file, sheet_name)
    
    def _import_excel_file_inner(self, excel_file, sheet_name: str = 'Cleaning schedule') -> Dict:
        """
        Import bookings from Excel file.
        
        Args:
            excel_file: Uploaded Excel file
            sheet_name: Name of the sheet to import (default: 'cleaning schedule')
            
        Returns:
            Dict with import results
        """
        try:
            # Read Excel file
            df = self._read_excel_file(excel_file, sheet_name)
            if df is None or df.empty:
                return {
                    'success': False,
                    'error': f'No data found in sheet "{sheet_name}" or file is empty'
                }
            
            self.total_rows = len(df)
            
            # Create import log
            self.import_log = self._create_import_log(excel_file)
            
            # First pass: identify all unique properties and handle new ones
            new_properties = self._identify_new_properties(df)
            
            # If there are new properties, handle them based on user role
            if new_properties:
                if not self.user.is_superuser:
                    # Non-admin users cannot proceed with new properties
                    return {
                        'success': False,
                        'requires_property_approval': True,
                        'new_properties': new_properties,
                        'message': f'Found {len(new_properties)} new properties that require admin approval. Please contact an administrator.',
                        'total_rows': self.total_rows
                    }
                else:
                    # Admin can create properties automatically
                    logger.info(f"Admin {self.user.username} will create {len(new_properties)} new properties during import")
                    # Continue with import - properties will be created as needed
            
            # Process each row individually (no transaction wrapper to allow partial success)
            processed_rows = 0
            successful_rows = 0
            skipped_rows = 0
            
            for index, row in df.iterrows():
                try:
                    # Skip completely empty rows
                    if row.isna().all():
                        logger.debug(f"Skipping empty row {index + 2}")
                        skipped_rows += 1
                        continue
                    
                    # Log row processing start
                    logger.debug(f"Processing row {index + 2}: {row.get('Guest name', 'No guest name')} - {row.get('Properties', 'No property')}")
                    
                    # Track success count before processing
                    success_before = self.success_count
                    
                    self._process_booking_row(row, index + 2)  # +2 because Excel is 1-indexed and has header
                    processed_rows += 1
                    
                    # Check if this row actually resulted in a booking change
                    if self.success_count > success_before:
                        successful_rows += 1
                        logger.debug(f"Row {index + 2}: Successfully processed (success count: {success_before} -> {self.success_count})")
                    else:
                        logger.warning(f"Row {index + 2}: Processed but no booking created/updated")
                    
                except Exception as e:
                    error_msg = f"Row {index + 2}: {str(e)}"
                    self.errors.append(error_msg)
                    logger.error(f"Error processing row {index + 2}: {e}")
                    continue
            
            logger.info(f"Row processing summary:")
            logger.info(f"  • Total rows in Excel: {self.total_rows}")
            logger.info(f"  • Rows processed: {processed_rows}")
            logger.info(f"  • Rows skipped (empty): {skipped_rows}")
            logger.info(f"  • Successful imports: {self.success_count}")
            logger.info(f"  • Rows with errors: {len(self.errors)}")
            
            # Update import log with results
            self._update_import_log()
            
            # Count new properties actually created during import
            new_properties_created = len([w for w in self.warnings if w.startswith('✨ Created new property:')])
            
            # Log the discrepancy for debugging
            if processed_rows != self.success_count:
                logger.warning(f"Row processing discrepancy: {processed_rows} rows processed, {self.success_count} successful imports")
                logger.warning(f"This suggests some rows were processed but didn't result in booking changes")
            
            return {
                'success': True,
                'total_rows': self.total_rows,
                'successful_imports': self.success_count,  # This is the actual count of created/updated bookings
                'processed_rows': processed_rows,  # This is the count of rows that were attempted
                'skipped_rows': skipped_rows,  # This is the count of rows that were skipped (empty)
                'errors_count': len(self.errors),
                'warnings_count': len(self.warnings),
                'errors': self.errors,
                'warnings': self.warnings,
                'import_log_id': self.import_log.id if self.import_log else None,
                'new_properties_created': new_properties_created,
                'new_properties_list': [w.replace('✨ Created new property: ', '') for w in self.warnings if w.startswith('✨ Created new property:')]
            }
            
        except Exception as e:
            logger.error(f"Excel import failed: {e}")
            return {
                'success': False,
                'error': f'Import failed: {str(e)}'
            }
    
    def _identify_new_properties(self, df) -> List[str]:
        """Identify new properties that don't exist in the database."""
        new_properties = []
        
        # Get all unique property names from Excel
        if 'Properties' in df.columns:
            excel_properties = df['Properties'].dropna().unique()
        else:
            # Try alternative column names
            for col in ['Property', 'Listing', 'Property Name']:
                if col in df.columns:
                    excel_properties = df[col].dropna().unique()
                    break
            else:
                return []
        
        # Check which properties don't exist in database
        for prop_name in excel_properties:
            if prop_name and str(prop_name).strip():
                prop_name = str(prop_name).strip()
                try:
                    # Check if property exists
                    Property.objects.get(name__iexact=prop_name)
                except Property.DoesNotExist:
                    # Try partial matches
                    if not Property.objects.filter(name__icontains=prop_name).exists():
                        new_properties.append(prop_name)
        
        return new_properties
    
    def _read_excel_file(self, excel_file, sheet_name: str):
        """Read Excel file and return DataFrame."""
        import pandas as pd  # Import pandas once at the top
        
        try:
            # Handle Django uploaded file properly
            if hasattr(excel_file, 'read'):
                # For Django uploaded files, we need to reset position and read as bytes
                excel_file.seek(0)
                file_content = excel_file.read()
                
                # Try to read with pandas first
                import io
                df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_name)
                return df
            else:
                # For regular file objects
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                return df
        except Exception as e:
            logger.warning(f"Pandas failed to read Excel: {e}, trying openpyxl")
            try:
                import openpyxl
                import io
                
                # Fallback to openpyxl with proper file handling
                if hasattr(excel_file, 'read'):
                    excel_file.seek(0)
                    file_content = excel_file.read()
                    workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
                else:
                    workbook = openpyxl.load_workbook(excel_file, read_only=True)
                    
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
                
                sheet = workbook[sheet_name]
                data = []
                headers = []
                
                for row in sheet.iter_rows(values_only=True):
                    if not headers:
                        headers = [str(cell).strip() if cell else '' for cell in row]
                    else:
                        data.append([str(cell).strip() if cell else '' for cell in row])
                
                # Now pd is properly available
                df = pd.DataFrame(data, columns=headers)
                workbook.close()
                return df
                
            except Exception as e2:
                logger.error(f"Both pandas and openpyxl failed: {e2}")
                raise ValueError(f"Could not read Excel file: {str(e2)}")
    
    def _create_import_log(self, excel_file) -> BookingImportLog:
        """Create import log entry."""
        # Create a default template if none exists
        if not self.template:
            # Get or create a default template for the first property
            from ..models import Property
            try:
                first_property = Property.objects.first()
                if first_property:
                    self.template, created = BookingImportTemplate.objects.get_or_create(
                        name="Default Import Template",
                        property_ref=first_property,
                        defaults={
                            'import_type': 'csv',
                            'auto_create_tasks': True,
                            'created_by': self.user
                        }
                    )
                else:
                    # Create a dummy template if no properties exist
                    self.template = BookingImportTemplate.objects.create(
                        name="Default Import Template",
                        property_ref=None,
                        import_type='csv',
                        auto_create_tasks=True,
                        created_by=self.user
                    )
            except Exception as e:
                logger.error(f"Could not create default template: {e}")
                # Create without template if all else fails
                return BookingImportLog.objects.create(
                    template=None,
                    import_file=excel_file,
                    total_rows=self.total_rows,
                    successful_imports=0,
                    errors_count=0,
                    errors_log='',
                    imported_by=self.user
                )
        
        return BookingImportLog.objects.create(
            template=self.template,
            import_file=excel_file,
            total_rows=self.total_rows,
            successful_imports=0,
            errors_count=0,
            errors_log='',
            imported_by=self.user
        )
    
    def _process_booking_row(self, row, row_number: int):
        """Process a single booking row from the Excel file."""
        # Use individual transaction for each row to allow partial success
        with transaction.atomic():
            try:
                # Extract and validate data
                booking_data = self._extract_booking_data(row, row_number)
                if not booking_data:
                    logger.warning(f"Row {row_number}: No booking data extracted - skipping")
                    return
                
                # Find or create property
                property_obj = self._find_or_create_property(booking_data['property_label_raw'])
                if not property_obj:
                    error_msg = f"Row {row_number}: Could not find or create property '{booking_data['property_label_raw']}'"
                    self.errors.append(error_msg)
                    logger.error(error_msg)
                    return
                
                # Check for booking conflicts
                conflicts = self._check_booking_conflicts(booking_data, property_obj)
                if conflicts:
                    for conflict in conflicts:
                        self.warnings.append(f"Row {row_number}: {conflict}")
                
                # Check for existing booking
                existing_booking = self._find_existing_booking(booking_data, property_obj)
                
                if existing_booking:
                    # Update existing booking
                    self._update_booking(existing_booking, booking_data, row)
                    
                    # Update associated task status if it exists
                    self._update_associated_task(existing_booking, booking_data)
                    
                    self.success_count += 1
                    logger.info(f"Updated booking {existing_booking.external_code} from row {row_number}")
                else:
                    # Create new booking
                    new_booking = self._create_booking(booking_data, property_obj, row)
                    self.success_count += 1
                    logger.info(f"Created new booking {new_booking.external_code} from row {row_number}")
                    
                    # Auto-create cleaning task if enabled
                    if self.template and self.template.auto_create_tasks:
                        self._create_cleaning_task(new_booking)
                        
            except Exception as e:
                error_msg = f"Row {row_number}: {str(e)}"
                self.errors.append(error_msg)
                logger.error(error_msg)
                # Re-raise to trigger transaction rollback for this row
                raise
    
    def _extract_booking_data(self, row, row_number: int) -> Optional[Dict]:
        """Extract and validate booking data from Excel row."""
        try:
            import pandas as pd
            # Map Excel columns to our fields
            data = {}
            
            # Required fields - based on actual Excel column names
            # Handle column name variations
            required_fields = {}
            
            # Map columns with fallbacks for variations
            column_mappings = {
                'external_code': ['Confirmation code'],
                'external_status': ['Status'],
                'guest_name': ['Guest name'],
                'guest_contact': ['Contact'],
                'source': ['Booking source', 'Airbnb/VRBO'],  # Handle both column names
                'listing_name': ['Listing'],
                'earnings_amount': ['Earnings'],
                'booked_on': ['Booked'],
                'adults': ['# of adults'],
                'children': ['# of children'],
                'infants': ['# of infants'],
                'start_date': ['Start date'],
                'end_date': ['End date'],
                'nights': ['# of nights'],
                'property_label_raw': ['Properties'],
                'same_day_note': ['Check ', 'Check 1']
            }
            
            # Find the first available column for each field
            for field_name, possible_columns in column_mappings.items():
                for col in possible_columns:
                    if col in row.index:
                        required_fields[col] = field_name
                        break
            
            # Debug: Log the actual column names found in the Excel file
            logger.debug(f"Excel columns found: {list(row.index)}")
            logger.debug(f"Looking for 'Properties' column: {'Properties' in row.index}")
            logger.debug(f"Column mappings: {required_fields}")
            
            # Extract data with validation
            for excel_col, field_name in required_fields.items():
                if excel_col in row.index:
                    value = row[excel_col]
                    if pd.isna(value) or value == '':
                        if field_name in ['external_code', 'guest_name', 'start_date', 'end_date']:
                            raise ValueError(f"Required field '{excel_col}' is empty")
                        continue
                    
                    # Clean and validate the value
                    cleaned_value = self._clean_field_value(field_name, value)
                    if cleaned_value is not None:
                        data[field_name] = cleaned_value
            
            # Handle same day notes - combine both check fields
            same_day_notes = []
            for col in ['Check ', 'Check 1']:
                if col in row.index and not pd.isna(row[col]) and str(row[col]).strip():
                    same_day_notes.append(str(row[col]).strip())
            
            if same_day_notes:
                data['same_day_note'] = ' | '.join(same_day_notes)
                data['same_day_flag'] = True
            
            # Generate confirmation code for Direct/Owner bookings if missing
            if not data.get('external_code'):
                source = data.get('source', '').lower()
                if 'direct' in source or 'owner' in source:
                    # Generate unique code for Direct/Owner bookings
                    base_code = 'Direct' if 'direct' in source else 'Owner Staying'
                    counter = 1
                    while True:
                        generated_code = f"{base_code} {counter:02d}"
                        # Check if this code already exists
                        if not Booking.objects.filter(external_code=generated_code).exists():
                            data['external_code'] = generated_code
                            logger.info(f"Generated confirmation code: {generated_code}")
                            break
                        counter += 1
                        if counter > 99:  # Safety limit
                            data['external_code'] = f"{base_code} {counter:02d}"
                            break
                else:
                    raise ValueError("Confirmation code is required for platform bookings")
            
            # Agent's recommendation: Remove global uniqueness check here
            # Will be handled in _create_booking/_update_booking with proper scope
            
            # Validate required fields
            if not data.get('start_date') or not data.get('end_date'):
                raise ValueError("Start date and end date are required")
            if not data.get('guest_name'):
                raise ValueError("Guest name is required")
            
            # Calculate nights if not provided or invalid
            if not data.get('nights') and data.get('start_date') and data.get('end_date'):
                try:
                    if isinstance(data['start_date'], datetime) and isinstance(data['end_date'], datetime):
                        nights = (data['end_date'] - data['start_date']).days
                        data['nights'] = max(1, nights)  # Ensure at least 1 night
                except Exception:
                    data['nights'] = 1  # Fallback
            
            return data
            
        except Exception as e:
            raise ValueError(f"Data extraction failed: {str(e)}")
    
    def _clean_field_value(self, field_name: str, value) -> any:
        """Clean and validate field values."""
        try:
            import pandas as pd
            if pd.isna(value) or value == '':
                return None
            
            # Handle pandas Timestamp objects first
            if hasattr(value, 'timestamp'):
                try:
                    if field_name in ['start_date', 'end_date', 'booked_on']:
                        # Convert pandas Timestamp to timezone-aware datetime
                        naive_datetime = value.to_pydatetime()
                        if timezone.is_naive(naive_datetime):
                            return timezone.make_aware(naive_datetime, timezone.get_current_timezone())
                        return naive_datetime
                    else:
                        return value.isoformat()
                except:
                    return str(value)
            
            # Handle pandas Timedelta objects
            if hasattr(value, 'total_seconds'):
                try:
                    return str(value)
                except:
                    return None
            
            # Convert to string first
            value_str = str(value).strip()
            
            if field_name in ['adults', 'children', 'infants']:
                # Handle numeric fields
                try:
                    return int(float(value_str))
                except (ValueError, TypeError):
                    return 0 if field_name in ['adults'] else 0
            
            elif field_name == 'nights':
                # Handle nights field - could be number or date
                try:
                    # First try to parse as number
                    return int(float(value_str))
                except (ValueError, TypeError):
                    # If it's a date or invalid value, return None
                    # This will be handled in data extraction to calculate from dates
                    return None
            
            elif field_name in ['earnings_amount']:
                # Handle decimal fields
                try:
                    return Decimal(str(value_str).replace('$', '').replace(',', ''))
                except (ValueError, TypeError):
                    return None
            
            elif field_name in ['start_date', 'end_date', 'booked_on']:
                # Handle date fields - make timezone-aware for Tampa, FL
                if isinstance(value, datetime):
                    # If it's already a datetime, make it timezone-aware
                    if timezone.is_naive(value):
                        return timezone.make_aware(value, timezone.get_current_timezone())
                    return value
                elif isinstance(value, str):
                    # Try common date formats
                    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                        try:
                            naive_datetime = datetime.strptime(value_str, fmt)
                            # Make timezone-aware for Tampa, FL (America/New_York)
                            return timezone.make_aware(naive_datetime, timezone.get_current_timezone())
                        except ValueError:
                            continue
                    raise ValueError(f"Could not parse date: {value_str}")
                else:
                    raise ValueError(f"Invalid date format: {value_str}")
            
            elif field_name in ['check_in_time', 'check_out_time']:
                # Handle time fields
                if isinstance(value, time):
                    return value
                elif isinstance(value, str):
                    try:
                        return datetime.strptime(value_str, '%H:%M').time()
                    except ValueError:
                        return None
                else:
                    return None
            
            elif field_name in ['source']:
                # Normalize booking source
                source_map = {
                    'airbnb': 'Airbnb',
                    'vrbo': 'VRBO', 
                    'booking.com': 'Booking.com',
                    'expedia': 'Expedia',
                    'direct': 'Direct',
                    'directly': 'Direct',
                    'owner staying': 'Owner Staying',  # Check longer phrases first
                    'owner': 'Owner'
                }
                value_lower = value_str.lower()
                for key, normalized in source_map.items():
                    if key in value_lower:
                        return normalized
                return value_str.title()
            
            else:
                # Default string fields
                return value_str
                
        except Exception as e:
            logger.warning(f"Error cleaning field {field_name}: {e}")
            return str(value) if value is not None else None
    
    def _find_or_create_property(self, property_label: str) -> Optional[Property]:
        """Find existing property or create new one based on label."""
        if not property_label:
            logger.warning("Property label is empty or None")
            return None
        
        logger.debug(f"Looking for property: '{property_label}'")
        
        # Try to find exact match first
        try:
            property_obj = Property.objects.get(name__iexact=property_label)
            logger.debug(f"Found exact property match: {property_obj.name}")
            return property_obj
        except Property.DoesNotExist:
            logger.debug(f"No exact match found for property: {property_label}")
            pass
        
        # Try partial matches
        try:
            property_obj = Property.objects.filter(name__icontains=property_label).first()
            if property_obj:
                logger.debug(f"Found partial property match: {property_obj.name}")
                return property_obj
        except Exception as e:
            logger.warning(f"Error during partial property search: {e}")
            pass
        
        # For new properties, handle based on user role
        if self.user.is_superuser:
            # Admin can create properties directly
            try:
                logger.info(f"Admin creating new property: {property_label}")
                property_obj = Property.objects.create(
                    name=property_label,
                    created_by=self.user,
                    modified_by=self.user
                )
                logger.info(f"Successfully created property: {property_obj.name}")
                
                # Add to warnings to highlight new property creation
                self.warnings.append(f"✨ Created new property: {property_label}")
                
                return property_obj
            except Exception as e:
                logger.error(f"Could not create property '{property_label}': {e}")
                return None
        elif hasattr(self.user, 'profile') and self.user.profile.role == 'manager':
            # Manager cannot create properties - this will be handled by the import process
            logger.warning(f"Manager encountered new property: {property_label} - will be queued for admin approval")
            return None
        else:
            # Regular users cannot create properties
            logger.warning(f"User {self.user.username} cannot create property: {property_label}")
            return None
    
    def _find_existing_booking(self, booking_data: Dict, property_obj: Property) -> Optional[Booking]:
        """Find existing booking by external code, property, and source for proper scoping."""
        if not booking_data.get('external_code'):
            return None
        
        try:
            # GPT Agent Fix: Use scoped booking lookup (property, source, external_code) instead of external_code alone
            source = booking_data.get('source', '').lower()
            return Booking.objects.get(
                property=property_obj,
                source=source,
                external_code=booking_data['external_code']
            )
        except Booking.DoesNotExist:
            # If not found by external code, try to find by guest name and dates
            if booking_data.get('guest_name') and booking_data.get('start_date'):
                try:
                    return Booking.objects.get(
                        guest_name=booking_data['guest_name'],
                        check_in_date__date=booking_data['start_date'].date(),
                        property=property_obj
                    )
                except Booking.DoesNotExist:
                    pass
            return None
    
    def _check_booking_conflicts(self, booking_data: Dict, property_obj: Property, existing_booking: Optional[Booking] = None) -> List[str]:
        """Check for booking conflicts and return list of conflict messages."""
        conflicts = []
        
        if not booking_data.get('start_date') or not booking_data.get('end_date'):
            return conflicts
        
        start_date = booking_data['start_date']
        end_date = booking_data['end_date']
        
        # Check for overlapping bookings on the same property
        overlapping_bookings = Booking.objects.filter(
            property=property_obj,
            check_in_date__lt=end_date,
            check_out_date__gt=start_date
        )
        
        if existing_booking:
            overlapping_bookings = overlapping_bookings.exclude(id=existing_booking.id)
        
        if overlapping_bookings.exists():
            conflict_details = []
            for conflict_booking in overlapping_bookings[:3]:  # Show first 3 conflicts
                conflict_details.append(
                    f"{conflict_booking.guest_name} ({conflict_booking.check_in_date.strftime('%Y-%m-%d')} to {conflict_booking.check_out_date.strftime('%Y-%m-%d')}"
                )
            conflicts.append(f"⚠️ Property conflict: Overlaps with {overlapping_bookings.count()} existing booking(s): {', '.join(conflict_details)}")
        
        # Check for same-day check-in/out conflicts across all properties
        same_day_checkins = Booking.objects.filter(
            check_in_date__date=start_date.date()
        ).exclude(property=property_obj)
        
        same_day_checkouts = Booking.objects.filter(
            check_out_date__date=end_date.date()
        ).exclude(property=property_obj)
        
        if same_day_checkins.exists() or same_day_checkouts.exists():
            total_same_day = same_day_checkins.count() + same_day_checkouts.count()
            conflicts.append(f"📅 Same-day conflict: {total_same_day} other booking(s) have check-in/out on the same date")
        
        return conflicts
    
    def _create_booking(self, booking_data: Dict, property_obj: Property, row) -> Booking:
        """Create new booking from Excel data."""
        # Ensure nights field has a valid value
        nights_value = booking_data.get('nights')
        if nights_value is None or not isinstance(nights_value, (int, float)):
            # Calculate nights from start/end dates
            try:
                if isinstance(booking_data['start_date'], datetime) and isinstance(booking_data['end_date'], datetime):
                    nights_value = (booking_data['end_date'] - booking_data['start_date']).days
                    nights_value = max(1, nights_value)  # Ensure at least 1 night
                else:
                    nights_value = 1  # Fallback
            except Exception:
                nights_value = 1  # Fallback
        
        # Create booking with all the new fields
        try:
            # Map Excel status to Django status
            django_status = self._map_excel_status_to_booking_status(booking_data.get('external_status'))
            
            # Agent's recommendation: External code de-dupe scope
            # Handle duplicate external codes scoped to (property, source, external_code)
            if booking_data.get('external_code'):
                original_code = booking_data['external_code']
                code = original_code
                src = booking_data.get('source', '')
                i = 1
                
                while Booking.objects.filter(
                    property=property_obj, 
                    source=src, 
                    external_code=code
                ).exists():
                    i += 1
                    code = f"{original_code} #{i}"
                
                if i > 1:
                    logger.warning(f"Duplicate external code '{original_code}' for property/source - generated unique code: '{code}'")
                    booking_data['external_code'] = code
            
            booking = Booking.objects.create(
                property=property_obj,
                check_in_date=booking_data['start_date'],
                check_out_date=booking_data['end_date'],
                guest_name=booking_data['guest_name'],
                guest_contact=booking_data.get('guest_contact', ''),
                status=django_status,  # Use mapped Django status
                external_code=booking_data['external_code'],
                external_status=booking_data.get('external_status', ''),
                source=booking_data.get('source', ''),
                listing_name=booking_data.get('listing_name', ''),
                earnings_amount=booking_data.get('earnings_amount'),
                earnings_currency='USD',
                booked_on=booking_data.get('booked_on'),  # Add the missing booked_on field
                adults=booking_data.get('adults', 1),
                children=booking_data.get('children', 0),
                infants=booking_data.get('infants', 0),
                nights=nights_value,
                check_in_time=booking_data.get('check_in_time'),
                check_out_time=booking_data.get('check_out_time'),
                property_label_raw=booking_data['property_label_raw'],
                same_day_note=booking_data.get('same_day_note', ''),
                same_day_flag=bool(booking_data.get('same_day_note')),
                raw_row=self._serialize_row_data(row)
            )
            logger.debug(f"Successfully created booking {booking.external_code} with status {django_status}")
            return booking
        except Exception as e:
            logger.error(f"Failed to create booking with external_code '{booking_data.get('external_code')}': {e}")
            raise
    
    def _update_booking(self, booking: Booking, booking_data: Dict, row):
        """Update existing booking with new data."""
        # Update fields that might have changed
        if 'guest_name' in booking_data:
            booking.guest_name = booking_data['guest_name']
        if 'guest_contact' in booking_data:
            booking.guest_contact = booking_data['guest_contact']
        if 'external_status' in booking_data:
            booking.external_status = booking_data['external_status']
            # Also update the Django status when external status changes
            django_status = self._map_excel_status_to_booking_status(booking_data['external_status'])
            if booking.status != django_status:
                old_status = booking.status
                booking.status = django_status
                logger.info(f"Updated booking {booking.external_code} status from '{old_status}' to '{django_status}'")
        if 'listing_name' in booking_data:
            booking.listing_name = booking_data['listing_name']
        if 'earnings_amount' in booking_data:
            booking.earnings_amount = booking_data['earnings_amount']
        if 'adults' in booking_data:
            booking.adults = booking_data['adults']
        if 'children' in booking_data:
            booking.children = booking_data['children']
        if 'infants' in booking_data:
            booking.infants = booking_data['infants']
        if 'nights' in booking_data:
            nights_value = booking_data['nights']
            if nights_value is not None and isinstance(nights_value, (int, float)):
                booking.nights = nights_value
            else:
                # Calculate nights from start/end dates if nights is invalid
                try:
                    if isinstance(booking_data.get('start_date'), datetime) and isinstance(booking_data.get('end_date'), datetime):
                        nights_value = (booking_data['end_date'] - booking_data['start_date']).days
                        booking.nights = max(1, nights_value)  # Ensure at least 1 night
                    elif isinstance(booking.check_in_date, datetime) and isinstance(booking.check_out_date, datetime):
                        nights_value = (booking.check_out_date - booking.check_in_date).days
                        booking.nights = max(1, nights_value)  # Ensure at least 1 night
                except Exception:
                    # Keep existing nights value if calculation fails
                    pass
        if 'same_day_note' in booking_data:
            booking.same_day_note = booking_data['same_day_note']
            booking.same_day_flag = bool(booking_data['same_day_note'])
        
        # Update booked_on if provided
        if 'booked_on' in booking_data:
            booking.booked_on = booking_data['booked_on']
        
        # Update dates if they changed
        if 'start_date' in booking_data:
            booking.check_in_date = booking_data['start_date']
        
        if 'end_date' in booking_data:
            booking.check_out_date = booking_data['end_date']
        
        # Update raw data and timestamps
        booking.raw_row = self._serialize_row_data(row)
        booking.last_import_update = timezone.now()
        
        try:
            booking.save()
            logger.debug(f"Successfully updated booking {booking.external_code}")
        except Exception as e:
            logger.error(f"Failed to save updated booking {booking.external_code}: {e}")
            raise
    
    def _combine_date_time(self, date_obj: datetime, time_obj: Optional[time]) -> datetime:
        """Combine date and time objects into datetime."""
        if isinstance(date_obj, datetime):
            if time_obj:
                combined = datetime.combine(date_obj.date(), time_obj)
            else:
                combined = date_obj
        elif isinstance(date_obj, datetime.date):
            if time_obj:
                combined = datetime.combine(date_obj, time_obj)
            else:
                combined = datetime.combine(date_obj, time.min)
        else:
            raise ValueError(f"Invalid date object: {date_obj}")
        
        # Make timezone-aware for Tampa, FL if it's naive
        if timezone.is_naive(combined):
            return timezone.make_aware(combined, timezone.get_current_timezone())
        return combined
    
    def _create_cleaning_task(self, booking: Booking):
        """Automatically create cleaning task for new booking."""
        try:
            # Calculate task due date (day before check-in)
            task_due_date = booking.check_in_date - timedelta(days=1)
            
            # Map Excel status to task status
            task_status = self._map_excel_status_to_task_status(booking.external_status)
            
            # Create cleaning task
            task = Task.objects.create(
                title=f"Pre-arrival Cleaning - {booking.property.name}",
                description=f"Cleaning for {booking.guest_name} arrival on {booking.check_in_date.strftime('%Y-%m-%d')} (Status: {booking.external_status})",
                task_type='cleaning',
                property=booking.property,
                booking=booking,
                status=task_status,
                due_date=task_due_date,
                created_by=self.user,
                modified_by=self.user
            )
            
            logger.info(f"Created cleaning task {task.id} for booking {booking.external_code} with status {task_status}")
            
        except Exception as e:
            logger.error(f"Failed to create cleaning task for booking {booking.external_code}: {e}")
    
    def _map_excel_status_to_booking_status(self, excel_status: str) -> str:
        """Map Excel status to appropriate Django booking status."""
        if not excel_status:
            return 'booked'
        
        excel_status_lower = excel_status.lower()
        
        # Map Excel statuses to Django booking statuses
        status_mapping = {
            'booked': 'booked',
            'confirmed': 'confirmed',
            'currently hosting': 'currently_hosting',
            'owner staying': 'owner_staying',
            'cancelled': 'cancelled',
            'canceled': 'cancelled',
            'completed': 'completed'
        }
        
        for excel_key, django_status in status_mapping.items():
            if excel_key in excel_status_lower:
                return django_status
        
        # Default to booked for unknown statuses
        return 'booked'
    
    def _map_excel_status_to_task_status(self, excel_status: str) -> str:
        """Map Excel status to appropriate task status."""
        if not excel_status:
            return 'pending'
        
        excel_status_lower = excel_status.lower()
        
        # Map Excel statuses to task statuses
        status_mapping = {
            'booked': 'pending',
            'confirmed': 'pending',
            'currently hosting': 'in-progress',
            'owner staying': 'pending',
            'cancelled': 'canceled',
            'canceled': 'canceled',
            'completed': 'completed'
        }
        
        for excel_key, task_status in status_mapping.items():
            if excel_key in excel_status_lower:
                return task_status
        
        # Default to pending for unknown statuses
        return 'pending'
    
    def _update_associated_task(self, booking: Booking, booking_data: Dict):
        """Update associated task status when booking status changes."""
        try:
            # Find associated task
            associated_task = Task.objects.filter(booking=booking).first()
            
            if associated_task and 'external_status' in booking_data:
                # Agent's recommendation: Respect lock mechanism for import protection
                if associated_task.is_locked_by_user:
                    logger.info(f"Skipping task {associated_task.id} update - locked by user")
                    return
                
                # Map new Excel status to task status
                new_task_status = self._map_excel_status_to_task_status(booking_data['external_status'])
                
                if associated_task.status != new_task_status:
                    old_status = associated_task.status
                    associated_task.status = new_task_status
                    associated_task.modified_by = self.user
                    associated_task.save()
                    
                    logger.info(f"Updated task {associated_task.id} status from '{old_status}' to '{new_task_status}' for booking {booking.external_code}")
                    
        except Exception as e:
            logger.error(f"Failed to update associated task for booking {booking.external_code}: {e}")
    
    def _serialize_row_data(self, row) -> Dict:
        """Convert pandas row data to JSON-serializable format."""
        import pandas as pd
        import json
        from datetime import datetime, date, time
        
        serialized = {}
        for key, value in row.items():
            if pd.isna(value):
                serialized[key] = None
            elif isinstance(value, datetime):
                serialized[key] = value.isoformat()
            elif isinstance(value, date):
                serialized[key] = value.isoformat()
            elif isinstance(value, time):
                serialized[key] = value.isoformat()
            elif hasattr(value, 'timestamp'):  # pandas Timestamp
                try:
                    serialized[key] = value.isoformat()
                except:
                    serialized[key] = str(value)
            elif hasattr(value, 'total_seconds'):  # pandas Timedelta
                serialized[key] = str(value)
            elif isinstance(value, (int, float, str, bool)):
                serialized[key] = value
            else:
                # Convert any other types to string
                serialized[key] = str(value)
        return serialized
    
    def _update_import_log(self):
        """Update import log with final results."""
        if self.import_log:
            self.import_log.successful_imports = self.success_count
            self.import_log.errors_count = len(self.errors)
            
            # Combine errors and warnings for comprehensive logging
            all_logs = []
            if self.errors:
                all_logs.append("=== ERRORS ===")
                all_logs.extend(self.errors)
            
            if self.warnings:
                all_logs.append("\n=== WARNINGS ===")
                all_logs.extend(self.warnings)
            
            if all_logs:
                self.import_log.errors_log = '\n'.join(all_logs)
            else:
                self.import_log.errors_log = "No errors or warnings"
            
            self.import_log.save()
            
            # Log summary to console
            logger.info(f"Import completed: {self.success_count}/{self.total_rows} successful, {len(self.errors)} errors, {len(self.warnings)} warnings")
            
            if self.warnings:
                logger.info("Warnings found during import:")
                for warning in self.warnings[:10]:  # Log first 10 warnings
                    logger.info(f"  - {warning}")
                if len(self.warnings) > 10:
                    logger.info(f"  ... and {len(self.warnings) - 10} more warnings")
